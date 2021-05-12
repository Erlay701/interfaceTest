import os
import sys
import time

import xlrd
import xlsxwriter
import xlutils
import xlwt
from xlsxwriter import Workbook
import getpathInfo
import shutil
from xlutils.copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import RED, GREEN, ORANGE

path = getpathInfo.get_Path()
xlsPath = os.path.join(path, "testFile", 'case', 'userCase.xlsx')
Path = os.path.join(path, "result")
now = time.strftime("%Y-%m-%d %H_%M_%S")

resultpath = Path + '/' + now + ' result.xlsx'


class WriteExcel:
    """文件写入数据"""

    def __init__(self):
        fileName = resultpath
        self.filename = fileName
        if not os.path.exists(self.filename):
            # 文件不存在，则拷贝模板文件至指定报告目录下
            shutil.copyfile(xlsPath, resultpath)

        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def get_rows(self):
        rows = self.ws.max_row
        return rows

    def write_data(self, name, msg, value):
        rows = self.get_rows()
        font_GREEN = Font(name='宋体', color=GREEN, bold=True)
        font_RED = Font(name='宋体', color=RED, bold=True)
        font1 = Font(name='宋体', color=ORANGE, bold=True)
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for i in range(1, rows + 1):
            F_N = "F" + str(i)
            G_N = "G" + str(i)
            if name == self.ws.cell(i, 1).value:
                if value == "pass":
                    self.ws.cell(i, 6, value)
                    self.ws.cell(i, 7, msg)
                    self.ws[F_N].font = font_GREEN
                    self.ws[F_N].alignment = align
                    self.ws[G_N].alignment = align
                if value == "fail":
                    self.ws.cell(i, 6, value)
                    self.ws.cell(i, 7, msg)
                    self.ws[F_N].font = font1
                    self.ws[F_N].alignment = align
                    self.ws[G_N].font = font1
                    self.ws[G_N].alignment = align
        self.wb.save(self.filename)


'''
    def write_data(self, row_n, value):
        """
        写入测试结果
        :param row_n:数据所在行数
        :param value: 测试结果值
        :return: 无
        """
        font_GREEN = Font(name='宋体', color=GREEN, bold=True)
        font_RED = Font(name='宋体', color=RED, bold=True)
        font1 = Font(name='宋体', color=DARKYELLOW, bold=True)
        align = Alignment(horizontal='center', vertical='center')
        # 获数所在行数
        L_n = "L" + str(row_n)
        M_n = "M" + str(row_n)

        if value == "PASS":
            self.ws.cell(row_n, 7, value)
            self.ws[L_n].font = font_GREEN
        if value == "FAIL":
            self.ws.cell(row_n, 7, value)
            self.ws[L_n].font = font_RED
        if value == "ERROR":
            self.ws.cell(row_n, 7, value)
            self.ws[L_n].font = font_RED
        self.ws[L_n].alignment = align
        self.ws[M_n].font = font1
        self.ws[M_n].alignment = align
        self.wb.save(self.filename)
'''

if __name__ == '__main__':
    WriteExcel()
